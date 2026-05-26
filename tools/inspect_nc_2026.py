"""First-look inspector for NC_2026.xlsx — what sheets, what columns, what range.

Run from worktree root:
    .venv/bin/python tools/inspect_nc_2026.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl


def main() -> None:
    path = Path("data/disclosures/NC_2026.xlsx")
    print(f"== File ==")
    print(f"  path: {path}")
    print(f"  size: {path.stat().st_size:,} bytes")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"\n== Sheets ({len(wb.sheetnames)}) ==")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  - {name!r}: {ws.max_row} rows x {ws.max_column} cols")

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n== Sheet: {name!r} ==")
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            print("  (empty)")
            continue
        print(f"  columns ({len(header)}):")
        for i, col in enumerate(header):
            print(f"    [{i:>2}] {col!r}")

        print(f"  first 3 data rows:")
        for n, row in enumerate(rows_iter):
            if n >= 3:
                break
            print(f"    row {n+2}: {row}")


if __name__ == "__main__":
    main()
