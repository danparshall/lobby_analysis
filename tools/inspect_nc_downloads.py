"""Inspect the 4 daily_*.xls files in ~/Downloads/lobby/nc/ (actually xlsx).

For each: sheets, row count, headers, first 2 data rows. Lets us infer what
each file represents before moving + renaming.

Run from worktree root:
    uv run python tools/inspect_nc_downloads.py
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl

FILES = [
    Path.home() / "Downloads/lobby/nc/daily_1.xls",
    Path.home() / "Downloads/lobby/nc/daily_2.xls",
    Path.home() / "Downloads/lobby/nc/daily-3.xls",
    Path.home() / "Downloads/lobby/nc/daily_4.xls",
]


def inspect(path: Path) -> None:
    print(f"\n{'=' * 70}")
    print(f"FILE: {path.name}  ({path.stat().st_size:,} bytes)")
    print(f"{'=' * 70}")
    # File extension is .xls but `file` reports XLSX content. Read via BytesIO
    # to bypass openpyxl's extension check.
    try:
        wb = openpyxl.load_workbook(
            BytesIO(path.read_bytes()), read_only=True, data_only=True
        )
    except Exception as exc:
        print(f"  ERROR opening: {exc}")
        return

    print(f"  sheets: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            print(f"\n  -- {sname!r}: EMPTY")
            continue
        data = list(rows_iter)
        print(f"\n  -- Sheet {sname!r}: {len(data):,} data rows, {len(header)} cols")
        print(f"     columns:")
        for i, col in enumerate(header):
            print(f"       [{i:>2}] {col!r}")
        print(f"     first 2 rows:")
        for n, row in enumerate(data[:2]):
            print(f"       row {n+2}: {row}")


def main() -> None:
    for p in FILES:
        if not p.exists():
            print(f"MISSING: {p}")
            continue
        inspect(p)


if __name__ == "__main__":
    main()
